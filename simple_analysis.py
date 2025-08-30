import openpyxl
import os

def analyze_excel_file():
    """
    Analyse simple du fichier Excel des données cliniques du cerveau
    """
    
    file_path = 'BrainClinicalData/MU-Glioma-Post_ClinicalData-July2025.xlsx'
    
    try:
        # Charger le workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Obtenir la première feuille
        sheet = workbook.active
        
        print("✅ Fichier Excel chargé avec succès!")
        print(f"📊 Nombre de lignes: {sheet.max_row}")
        print(f"📋 Nombre de colonnes: {sheet.max_column}")
        
        # Lire les en-têtes (première ligne)
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value)
        
        print(f"\n📝 Colonnes disponibles ({len(headers)}):")
        for i, header in enumerate(headers, 1):
            print(f"{i:2d}. {header}")
        
        # Analyser quelques lignes de données
        print(f"\n👀 Aperçu des données (5 premières lignes):")
        for row in range(2, min(7, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "None")
            print(f"Ligne {row-1}: {row_data[:5]}...")  # Afficher seulement les 5 premières colonnes
        
        # Compter les valeurs non-nulles par colonne
        print(f"\n📊 Analyse des valeurs non-nulles par colonne:")
        for col in range(1, min(11, sheet.max_column + 1)):  # Analyser les 10 premières colonnes
            non_null_count = 0
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    non_null_count += 1
            
            header = headers[col-1] if col <= len(headers) else f"Colonne_{col}"
            print(f"{header}: {non_null_count}/{sheet.max_row-1} valeurs non-nulles")
        
        # Rechercher des variables cibles potentielles
        print(f"\n🎯 Recherche de variables cibles potentielles:")
        target_keywords = ['survival', 'mortality', 'death', 'outcome', 'status', 'prognosis', 
                          'recurrence', 'progression', 'response', 'grade', 'stage', 'type']
        
        potential_targets = []
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                for keyword in target_keywords:
                    if keyword in header_lower:
                        potential_targets.append(header)
                        break
        
        if potential_targets:
            print("Variables cibles potentielles trouvées:")
            for target in potential_targets:
                print(f"  - {target}")
        else:
            print("Aucune variable cible évidente trouvée dans les noms de colonnes.")
        
        workbook.close()
        return headers
        
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")
        return None

if __name__ == "__main__":
    print("🧠 Analyse simple des données cliniques du cerveau")
    print("=" * 50)
    
    headers = analyze_excel_file()
    
    if headers:
        print(f"\n✅ Analyse terminée avec succès!")
        print(f"📁 {len(headers)} colonnes identifiées dans le fichier.")
    else:
        print("\n❌ Échec de l'analyse des données.")
