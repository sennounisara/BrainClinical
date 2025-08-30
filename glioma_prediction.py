import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
import joblib
import streamlit as st
import pandas as pd

def load_and_preprocess_data():
    """
    Charge et prétraite les données cliniques des gliomes
    """
    
    file_path = 'BrainClinicalData/MU-Glioma-Post_ClinicalData-July2025.xlsx'
    
    try:
        # Charger le workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Lire les en-têtes
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value)
        
        # Lire les données
        data = []
        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        # Créer un DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        workbook.close()
        
        print(f"✅ Données chargées: {df.shape}")
        return df
        
    except Exception as e:
        print(f"❌ Erreur lors du chargement: {e}")
        return None

def prepare_features(df):
    """
    Prépare les features pour la prédiction
    """
    
    # Sélectionner les variables pertinentes pour la prédiction
    feature_columns = [
        'Sex at Birth',
        'Age at diagnosis',
        'Primary Diagnosis',
        'Grade of Primary Brain Tumor',
        'Stereotactic Biopsy before Surgical Resection',
        'IDH1 mutation',
        'IDH2 mutation',
        '1p/19q',
        'ATRX mutation',
        'MGMT methylation',
        'BRAF V600E mutation',
        'TERT promoter mutation',
        'Chromosome 7 gain and Chromosome 10 loss',
        'H3-3A mutation',
        'EGFR amplification',
        'PTEN mutation',
        'CDKN2A/B deletion',
        'TP53 alteration',
        'Previous Brain Tumor',
        'Initial Chemo Therapy',
        'Radiation Therapy'
    ]
    
    # Variables cibles potentielles
    target_columns = [
        'Progression',
        'Overall Survival (Death)',
        'Grade of Primary Brain Tumor'
    ]
    
    # Filtrer les colonnes disponibles
    available_features = [col for col in feature_columns if col in df.columns]
    available_targets = [col for col in target_columns if col in df.columns]
    
    print(f"📊 Features disponibles: {len(available_features)}")
    print(f"🎯 Variables cibles disponibles: {len(available_targets)}")
    
    # Créer un sous-ensemble avec les colonnes disponibles
    selected_df = df[available_features + available_targets].copy()
    
    # Supprimer les lignes avec trop de valeurs manquantes
    threshold = len(available_features) * 0.5  # Au moins 50% des features doivent être présentes
    selected_df = selected_df.dropna(thresh=threshold)
    
    print(f"📈 Données après nettoyage: {selected_df.shape}")
    
    return selected_df, available_features, available_targets

def encode_categorical_features(df, feature_columns):
    """
    Encode les variables catégorielles
    """
    
    df_encoded = df.copy()
    encoders = {}
    
    for col in feature_columns:
        if df_encoded[col].dtype == 'object':
            # Créer un encodeur pour cette colonne
            le = LabelEncoder()
            
            # Remplacer les valeurs manquantes par une valeur spéciale
            df_encoded[col] = df_encoded[col].fillna('Unknown')
            
            # Encoder les valeurs
            df_encoded[col] = le.fit_transform(df_encoded[col].astype(str))
            encoders[col] = le
    
    return df_encoded, encoders

def train_models(df, feature_columns, target_columns):
    """
    Entraîne les modèles de prédiction
    """
    
    models = {}
    scalers = {}
    encoders = {}
    
    for target in target_columns:
        if target in df.columns:
            print(f"\n🎯 Entraînement du modèle pour: {target}")
            
            # Préparer les données pour cette cible
            target_df = df[feature_columns + [target]].dropna()
            
            if len(target_df) < 10:  # Trop peu de données
                print(f"⚠️ Pas assez de données pour {target}")
                continue
            
            # Encoder les features catégorielles
            target_df_encoded, feature_encoders = encode_categorical_features(target_df, feature_columns)
            
            # Encoder la variable cible si elle est catégorielle
            if target_df_encoded[target].dtype == 'object':
                target_encoder = LabelEncoder()
                target_df_encoded[target] = target_encoder.fit_transform(target_df_encoded[target].astype(str))
                encoders[target] = target_encoder
            else:
                encoders[target] = None
            
            # Séparer features et cible
            X = target_df_encoded[feature_columns]
            y = target_df_encoded[target]
            
            # Diviser en train/test
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
            
            # Standardiser les features
            scaler = StandardScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_test_scaled = scaler.transform(X_test)
            
            # Entraîner Random Forest
            rf_model = RandomForestClassifier(n_estimators=100, random_state=42)
            rf_model.fit(X_train_scaled, y_train)
            
            # Évaluer le modèle
            y_pred = rf_model.predict(X_test_scaled)
            accuracy = accuracy_score(y_test, y_pred)
            
            print(f"📊 Précision pour {target}: {accuracy:.3f}")
            
            # Sauvegarder le modèle
            models[target] = rf_model
            scalers[target] = scaler
    
    return models, scalers, encoders

def create_prediction_app(models, scalers, encoders, feature_columns):
    """
    Crée l'application Streamlit pour les prédictions
    """
    
    st.title('🧠 Prédiction des Gliomes - Analyse Clinique')
    
    st.sidebar.title('Informations sur l\'Application')
    st.sidebar.markdown("""
    Cette application utilise des modèles de Machine Learning pour prédire 
    différents aspects des gliomes basés sur les données cliniques.
    
    ### Variables prédites:
    - **Progression**: Risque de progression de la tumeur
    - **Survie globale**: Prédiction de survie
    - **Grade de la tumeur**: Classification du grade tumoral
    
    ### Facteurs pris en compte:
    - Caractéristiques démographiques (âge, sexe)
    - Mutations génétiques (IDH1, IDH2, MGMT, etc.)
    - Antécédents médicaux
    - Traitements reçus
    """)
    
    st.header('📋 Saisie des Données Patient')
    
    # Formulaire de saisie
    with st.form("prediction_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            sex = st.selectbox('Sexe à la naissance', ['Male', 'Female', 'Unknown'])
            age = st.number_input('Âge au diagnostic', min_value=0, max_value=100, value=50)
            primary_diagnosis = st.selectbox('Diagnostic primaire', 
                                           ['Glioblastoma', 'Astrocytoma', 'Oligodendroglioma', 'Unknown'])
            grade = st.selectbox('Grade de la tumeur primaire', 
                               ['Grade I', 'Grade II', 'Grade III', 'Grade IV', 'Unknown'])
            biopsy = st.selectbox('Biopsie stéréotaxique avant résection', 
                                ['Yes', 'No', 'Unknown'])
        
        with col2:
            idh1 = st.selectbox('Mutation IDH1', ['Positive', 'Negative', 'Unknown'])
            idh2 = st.selectbox('Mutation IDH2', ['Positive', 'Negative', 'Unknown'])
            codeletion = st.selectbox('Codeletion 1p/19q', ['Present', 'Absent', 'Unknown'])
            mgmt = st.selectbox('Méthylation MGMT', ['Methylated', 'Unmethylated', 'Unknown'])
            egfr = st.selectbox('Amplification EGFR', ['Present', 'Absent', 'Unknown'])
        
        # Traitements
        st.subheader('💊 Traitements')
        chemo = st.selectbox('Chimiothérapie initiale', ['Yes', 'No', 'Unknown'])
        radiation = st.selectbox('Radiothérapie', ['Yes', 'No', 'Unknown'])
        previous_tumor = st.selectbox('Tumeur cérébrale antérieure', ['Yes', 'No', 'Unknown'])
        
        submitted = st.form_submit_button("🔮 Faire la Prédiction")
    
    if submitted:
        st.header('🎯 Résultats de la Prédiction')
        
        # Préparer les données d'entrée
        input_data = {
            'Sex at Birth': sex,
            'Age at diagnosis': age,
            'Primary Diagnosis': primary_diagnosis,
            'Grade of Primary Brain Tumor': grade,
            'Stereotactic Biopsy before Surgical Resection': biopsy,
            'IDH1 mutation': idh1,
            'IDH2 mutation': idh2,
            '1p/19q': codeletion,
            'MGMT methylation': mgmt,
            'EGFR amplification': egfr,
            'Previous Brain Tumor': previous_tumor,
            'Initial Chemo Therapy': chemo,
            'Radiation Therapy': radiation
        }
        
        # Faire les prédictions pour chaque modèle
        for target, model in models.items():
            if target in scalers and target in encoders:
                st.subheader(f'📊 Prédiction: {target}')
                
                # Préparer les données d'entrée
                input_features = []
                for col in feature_columns:
                    if col in input_data:
                        value = input_data[col]
                        # Encoder si nécessaire
                        if col in encoders[target]:
                            try:
                                encoded_value = encoders[target][col].transform([str(value)])[0]
                            except:
                                encoded_value = 0  # Valeur par défaut
                        else:
                            encoded_value = value if isinstance(value, (int, float)) else 0
                        input_features.append(encoded_value)
                    else:
                        input_features.append(0)
                
                # Standardiser et prédire
                input_scaled = scalers[target].transform([input_features])
                prediction = model.predict(input_scaled)[0]
                probabilities = model.predict_proba(input_scaled)[0]
                
                # Afficher les résultats
                if target == 'Progression':
                    if prediction == 1:
                        st.warning("⚠️ Risque de progression détecté")
                    else:
                        st.success("✅ Faible risque de progression")
                
                elif target == 'Overall Survival (Death)':
                    if prediction == 1:
                        st.error("💀 Risque de mortalité élevé")
                    else:
                        st.success("✅ Bon pronostic de survie")
                
                elif target == 'Grade of Primary Brain Tumor':
                    grade_labels = ['Grade I', 'Grade II', 'Grade III', 'Grade IV']
                    if prediction < len(grade_labels):
                        predicted_grade = grade_labels[prediction]
                        st.info(f"📋 Grade prédit: {predicted_grade}")
                
                # Afficher la confiance
                max_prob = max(probabilities)
                st.metric("Confiance du modèle", f"{max_prob:.1%}")

def main():
    """
    Fonction principale
    """
    
    print("🧠 Démarrage de l'analyse des gliomes...")
    
    # Charger les données
    df = load_and_preprocess_data()
    if df is None:
        return
    
    # Préparer les features
    df_processed, feature_columns, target_columns = prepare_features(df)
    
    # Entraîner les modèles
    models, scalers, encoders = train_models(df_processed, feature_columns, target_columns)
    
    # Sauvegarder les modèles
    joblib.dump(models, 'glioma_models.pkl')
    joblib.dump(scalers, 'glioma_scalers.pkl')
    joblib.dump(encoders, 'glioma_encoders.pkl')
    joblib.dump(feature_columns, 'glioma_features.pkl')
    
    print("✅ Modèles entraînés et sauvegardés!")
    
    # Créer l'application Streamlit
    create_prediction_app(models, scalers, encoders, feature_columns)

if __name__ == "__main__":
    main()
