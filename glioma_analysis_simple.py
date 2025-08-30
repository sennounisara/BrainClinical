import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score
import joblib

def load_data():
    """
    Charge les données cliniques des gliomes
    """
    
    file_path = 'BrainClinicalData/MU-Glioma-Post_ClinicalData-July2025.xlsx'
    
    try:
        # Charger le workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Lire les en-têtes
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value)
        
        # Lire les données
        data = []
        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        workbook.close()
        
        print(f"✅ Données chargées: {len(data)} patients, {len(headers)} variables")
        return data, headers
        
    except Exception as e:
        print(f"❌ Erreur lors du chargement: {e}")
        return None, None

def prepare_features(data, headers):
    """
    Prépare les features pour la prédiction
    """
    
    # Variables d'intérêt pour la prédiction
    feature_columns = [
        'Sex at Birth',
        'Age at diagnosis',
        'Primary Diagnosis',
        'Grade of Primary Brain Tumor',
        'IDH1 mutation',
        'IDH2 mutation',
        '1p/19q',
        'MGMT methylation',
        'EGFR amplification',
        'Previous Brain Tumor',
        'Initial Chemo Therapy',
        'Radiation Therapy'
    ]
    
    # Variables cibles
    target_columns = [
        'Progression',
        'Overall Survival (Death)',
        'Grade of Primary Brain Tumor'
    ]
    
    # Trouver les indices des colonnes
    feature_indices = []
    for col in feature_columns:
        if col in headers:
            feature_indices.append(headers.index(col))
    
    target_indices = []
    for col in target_columns:
        if col in headers:
            target_indices.append(headers.index(col))
    
    print(f"📊 Features trouvées: {len(feature_indices)}")
    print(f"🎯 Variables cibles trouvées: {len(target_indices)}")
    
    # Extraire les données pertinentes
    X_data = []
    y_data = {}
    
    for row in data:
        # Vérifier si la ligne a suffisamment de données
        valid_features = 0
        feature_values = []
        
        for idx in feature_indices:
            value = row[idx] if idx < len(row) else None
            if value is not None and str(value).strip() != '':
                valid_features += 1
            feature_values.append(value)
        
        # Si au moins 50% des features sont présentes
        if valid_features >= len(feature_indices) * 0.5:
            X_data.append(feature_values)
            
            # Extraire les variables cibles
            for i, target_idx in enumerate(target_indices):
                target_name = target_columns[i]
                if target_name not in y_data:
                    y_data[target_name] = []
                
                target_value = row[target_idx] if target_idx < len(row) else None
                y_data[target_name].append(target_value)
    
    print(f"📈 Données valides: {len(X_data)} patients")
    
    return X_data, y_data, feature_columns[:len(feature_indices)], target_columns[:len(target_indices)]

def encode_categorical_data(X_data, y_data, feature_names, target_names):
    """
    Encode les données catégorielles
    """
    
    # Encoder les features
    X_encoded = []
    feature_encoders = {}
    
    for i, feature_name in enumerate(feature_names):
        # Extraire les valeurs pour cette feature
        feature_values = [row[i] for row in X_data]
        
        # Créer un encodeur
        le = LabelEncoder()
        
        # Remplacer les valeurs manquantes
        clean_values = []
        for val in feature_values:
            if val is None or str(val).strip() == '':
                clean_values.append('Unknown')
            else:
                clean_values.append(str(val))
        
        # Encoder
        try:
            encoded_values = le.fit_transform(clean_values)
            feature_encoders[feature_name] = le
        except:
            # Si l'encodage échoue, utiliser des valeurs numériques simples
            unique_values = list(set(clean_values))
            value_map = {val: idx for idx, val in enumerate(unique_values)}
            encoded_values = [value_map[val] for val in clean_values]
            feature_encoders[feature_name] = value_map
        
        # Ajouter à X_encoded
        if len(X_encoded) == 0:
            X_encoded = [[val] for val in encoded_values]
        else:
            for j, val in enumerate(encoded_values):
                X_encoded[j].append(val)
    
    # Encoder les variables cibles
    y_encoded = {}
    target_encoders = {}
    
    for target_name in target_names:
        if target_name in y_data:
            target_values = y_data[target_name]
            
            # Nettoyer les valeurs
            clean_target_values = []
            for val in target_values:
                if val is None or str(val).strip() == '':
                    clean_target_values.append('Unknown')
                else:
                    clean_target_values.append(str(val))
            
            # Encoder
            le = LabelEncoder()
            try:
                encoded_target_values = le.fit_transform(clean_target_values)
                target_encoders[target_name] = le
            except:
                unique_values = list(set(clean_target_values))
                value_map = {val: idx for idx, val in enumerate(unique_values)}
                encoded_target_values = [value_map[val] for val in clean_target_values]
                target_encoders[target_name] = value_map
            
            y_encoded[target_name] = encoded_target_values
    
    return X_encoded, y_encoded, feature_encoders, target_encoders

def train_models(X_encoded, y_encoded, feature_names, target_names):
    """
    Entraîne les modèles de prédiction
    """
    
    models = {}
    scalers = {}
    
    for target_name in target_names:
        if target_name in y_encoded:
            print(f"\n🎯 Entraînement du modèle pour: {target_name}")
            
            # Préparer les données
            X = np.array(X_encoded)
            y = np.array(y_encoded[target_name])
            
            # Supprimer les lignes avec des valeurs manquantes dans la cible
            valid_indices = ~np.isnan(y)
            X_valid = X[valid_indices]
            y_valid = y[valid_indices]
            
            if len(X_valid) < 10:
                print(f"⚠️ Pas assez de données pour {target_name}")
                continue
            
            # Diviser en train/test
            X_train, X_test, y_train, y_test = train_test_split(
                X_valid, y_valid, test_size=0.2, random_state=42
            )
            
            # Standardiser
            scaler = StandardScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_test_scaled = scaler.transform(X_test)
            
            # Entraîner Random Forest
            rf_model = RandomForestClassifier(n_estimators=100, random_state=42)
            rf_model.fit(X_train_scaled, y_train)
            
            # Évaluer
            y_pred = rf_model.predict(X_test_scaled)
            accuracy = accuracy_score(y_test, y_pred)
            
            print(f"📊 Précision pour {target_name}: {accuracy:.3f}")
            
            # Sauvegarder
            models[target_name] = rf_model
            scalers[target_name] = scaler
    
    return models, scalers

def main():
    """
    Fonction principale
    """
    
    print("🧠 Analyse des données cliniques des gliomes")
    print("=" * 50)
    
    # Charger les données
    data, headers = load_data()
    if data is None:
        return
    
    # Préparer les features
    X_data, y_data, feature_names, target_names = prepare_features(data, headers)
    
    if len(X_data) == 0:
        print("❌ Aucune donnée valide trouvée")
        return
    
    # Encoder les données
    X_encoded, y_encoded, feature_encoders, target_encoders = encode_categorical_data(
        X_data, y_data, feature_names, target_names
    )
    
    # Entraîner les modèles
    models, scalers = train_models(X_encoded, y_encoded, feature_names, target_names)
    
    # Sauvegarder les modèles
    joblib.dump(models, 'glioma_models.pkl')
    joblib.dump(scalers, 'glioma_scalers.pkl')
    joblib.dump(feature_encoders, 'glioma_feature_encoders.pkl')
    joblib.dump(target_encoders, 'glioma_target_encoders.pkl')
    joblib.dump(feature_names, 'glioma_feature_names.pkl')
    
    print("\n✅ Modèles entraînés et sauvegardés!")
    print("📁 Fichiers créés:")
    print("  - glioma_models.pkl")
    print("  - glioma_scalers.pkl")
    print("  - glioma_feature_encoders.pkl")
    print("  - glioma_target_encoders.pkl")
    print("  - glioma_feature_names.pkl")

if __name__ == "__main__":
    main()
